"""modules imports
"""
import os
import uuid
import openpyxl
from dotenv import load_dotenv
import pandas as pd
from tools.log_helper import log_helper

class file_helper:
    """functions to manipulate files
    """
    def __init__(self):
        load_dotenv()
        self.output_folder: str = str(os.getenv('OUTPUT_FOLDER'))\
            if os.getenv('OUTPUT_FOLDER') is not None else "./"
        self.log_manager = log_helper()

    def read_file_to_dataframe(
        self,
        file_path: str,
        file_extension: str,
        worksheet_name: str = "",
        separator: str = ",",
        encoding: str = "",
        xlsx_header: int = 0):
        """files to dataframe
        """
        if file_extension == "csv":
            if encoding == "":
                return pd.read_csv(file_path, sep=separator)
            return pd.read_csv(file_path, sep=separator, encoding=encoding)
        elif file_extension == "xlsx":
            if worksheet_name == "":
                return pd.read_excel(file_path)
            return pd.read_excel(file_path, worksheet_name, header=xlsx_header)
        else:
            raise BaseException(
                        "File extension error : " + file_extension + " is not recognized. ")\
                        from None

    def write_csv(self, header: str, content: list):
        """csv file writer
        """
        file_content = ""

        for record in content:
            file_content = file_content + (";".join(map(str, record))) + "\n"
        file_name = str(uuid.uuid1())
        with open(self.output_folder + "/" + file_name + ".csv", "a", encoding="utf-8") as file:
            file.write(header)
            file.write(file_content)

            return file

    def write_dataframe_to_csv(self, df_data: pd.DataFrame):
        """csv file writer
        """
        file_name = str(uuid.uuid1())
        df_data.to_csv(
            self.output_folder + "/" + file_name + ".csv",
            index=False,
            sep=";")

        return self.output_folder + "/" + file_name + ".csv"

    def write_xlsx(self, names: list, headers: list[list], contents: list[list[list]]):
        """xlsx file writer
        """
        file_name = str(uuid.uuid1())

        wbk = openpyxl.Workbook()
        # Removes the default 'Sheet1'
        del wbk[wbk.sheetnames[0]]

        for idx, content in enumerate(contents):
            wst = wbk.create_sheet(names[idx])
            wst.append(headers[idx])
            for row in content:
                print(row)
                print(type(row))
                wst.append(row)

        wbk.save(self.output_folder + '/' + f'{file_name}.xlsx')

        return self.output_folder + '/' + f'{file_name}.xlsx'

    def write_file(self, file_content: str, file_extension: str):
        """generic file writer
        """
        file_name = str(uuid.uuid1())
        with open(
                    self.output_folder + '/' + file_name + "." + file_extension\
                    , "a"
                    , encoding="utf-8"
        ) as file:
            file.write(file_content)

        return file
